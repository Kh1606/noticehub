"""Build src/data/regions.json from version2.xlsx.

Layout of version2.xlsx (verified):
  - Row 1: blank
  - Row 2: header (번호, 지역, 기관, 페이지명, 링크, 비고, crawlable)
  - Row 3+: data. Columns:
      A = number
      B = region (merged-cell: forward-fill when blank)
      C = sub-entity name (merged-cell: forward-fill when blank)
      D = page name
      E = link cell — display text "바로가기", real URL on cell.hyperlink.target
      F = notes
      G = crawlable (bool)

Output shape (mirrors the existing clt-plus regions.json):
  [
    { "region": "...",
      "subEntities": [
        { "name": "...",
          "sources": [ {"page": "...", "url": "...", "sheet": "...", "crawlable": false}, ... ] }
      ] }
  ]

Run from repo root:
    python scripts/build_regions_json.py
"""
from collections import OrderedDict
from pathlib import Path
import json
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "version2.xlsx"
OUT = ROOT / "src" / "data" / "regions.json"

HEADER_ROW = 2
COL_REGION = 2
COL_SUBENTITY = 3
COL_PAGE = 4
COL_LINK = 5
COL_CRAWLABLE = 7

REGION_FALLBACK = "-"
SUBENTITY_FALLBACK = "-"
PAGE_FALLBACK = "공지사항"

def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None

def main():
    wb = load_workbook(XLSX)
    # OrderedDict[region] -> OrderedDict[subEntity name] -> list of sources
    grouped = OrderedDict()
    total_sources = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        last_region = None
        last_sub = None
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            link_cell = ws.cell(row=r, column=COL_LINK)
            if not link_cell.hyperlink:
                continue
            url = link_cell.hyperlink.target
            if not url:
                continue

            region = _txt(ws.cell(row=r, column=COL_REGION).value) or last_region or REGION_FALLBACK
            sub = _txt(ws.cell(row=r, column=COL_SUBENTITY).value) or last_sub or SUBENTITY_FALLBACK
            page = _txt(ws.cell(row=r, column=COL_PAGE).value) or PAGE_FALLBACK
            crawl_val = ws.cell(row=r, column=COL_CRAWLABLE).value
            crawlable = bool(crawl_val) if crawl_val is not None else False

            last_region = region
            last_sub = sub

            grouped.setdefault(region, OrderedDict()).setdefault(sub, []).append(
                {"page": page, "url": url, "sheet": sn, "crawlable": crawlable}
            )
            total_sources += 1

    result = [
        {
            "region": region,
            "subEntities": [{"name": name, "sources": sources} for name, sources in subs.items()],
        }
        for region, subs in grouped.items()
    ]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} - {len(result)} regions, {total_sources} sources total")

if __name__ == "__main__":
    main()
