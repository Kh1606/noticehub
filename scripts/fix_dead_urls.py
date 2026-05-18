"""Patch version2.xlsx to replace 5 dead/wrong URLs with working ones found
by probing each site's homepage and validating against existing helpers.

OLD → NEW mapping (verified locally with the listed helper):
  - 평택시 saeol gosiList.do (404)
      → /pyeongtaek/saeol/gosi/list.do?seCode=01&mid=0401020000   (scrape_saeol, 10)
  - 안성시 portal/saeol/gosiList.do?mId=... (redirect to /main)
      → /portal/saeol/gosiList.do?mId=0501040000                  (scrape_saeol, 10)
  - 안산시 selectContents.do?cntnts_id=C0001209 (error.do)
      → /www/common/bbs/selectPageListBbs.do?bbs_code=WWW13       (scrape_fn_go_detail, 15)
  - 인천도시공사 입찰 open_content/.../bid.jsp (500)
      → /main/sale_lease/board/shopping_notice.jsp                (simple_list req=bbsMsgDetail, 11)
  - 한국수자원공사(물산업플랫폼) wis/wq/index.do (WebSquare/JS)
      → /news/sub01/noti01List.do?s_mid=105&brdId=KO27            (simple_table title_col=3, 10)

의왕시 URL is unchanged (the original works, just needs the eminwon helper).
Hyperlink target updates are applied to the link cell (col E) so the
rendered "바로가기" text stays the same.

Re-runnable: matching keys is exact-string on the OLD URL; safe if already
patched (the if-match check makes it a no-op).
"""
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "version2.xlsx"

OLD_TO_NEW = {
    # 평택시
    "https://www.pyeongtaek.go.kr/pyeongtaek/saeol/gosiList.do?seCode=01&mId=0401020000":
        "https://www.pyeongtaek.go.kr/pyeongtaek/saeol/gosi/list.do?seCode=01&mid=0401020000",
    # 안성시 (wrong mId: 0401040000 redirects to /main; 0501040000 hits the gosi list)
    "https://www.anseong.go.kr/portal/saeol/gosiList.do?mId=0401040000":
        "https://www.anseong.go.kr/portal/saeol/gosiList.do?mId=0501040000",
    # 안산시 (selectContents → selectPageListBbs)
    "https://www.ansan.go.kr/www/common/cntnts/selectContents.do?cntnts_id=C0001209":
        "https://www.ansan.go.kr/www/common/bbs/selectPageListBbs.do?bbs_code=WWW13",
    # 인천도시공사 입찰
    "https://www.ih.co.kr/open_content/main/customer/notification/bid.jsp":
        "https://www.ih.co.kr/main/sale_lease/board/shopping_notice.jsp",
    # 한국수자원공사(물산업플랫폼) — switch to the actual notice board
    "https://www.kwater.or.kr/wis/wq/index.do?w2xPath=/wis/ui/index.xml":
        "https://www.kwater.or.kr/news/sub01/noti01List.do?s_mid=105&brdId=KO27",
}


def main() -> None:
    wb = load_workbook(XLSX)
    patched = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                if not cell.hyperlink:
                    continue
                tgt = cell.hyperlink.target or ""
                if tgt in OLD_TO_NEW:
                    new = OLD_TO_NEW[tgt]
                    cell.hyperlink.target = new
                    print(f"  {sn} row {cell.row}: {tgt[:60]}... -> {new[:60]}...")
                    patched += 1
    wb.save(XLSX)
    print(f"\nPatched {patched} hyperlinks (expected 4).")


if __name__ == "__main__":
    main()
