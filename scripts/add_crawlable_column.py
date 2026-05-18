"""Add a `crawlable` boolean column (G) to version2.xlsx, defaulting every
data row to FALSE. Idempotent — re-running overwrites the same cells.

Run from repo root:
    python scripts/add_crawlable_column.py
"""
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "version2.xlsx"
HEADER_ROW = 2
LINK_COL = 5
CRAWLABLE_COL = 7

def main():
    wb = load_workbook(XLSX)
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws.cell(row=HEADER_ROW, column=CRAWLABLE_COL).value = "crawlable"
        marked = 0
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            link_cell = ws.cell(row=r, column=LINK_COL)
            if link_cell.hyperlink:
                ws.cell(row=r, column=CRAWLABLE_COL).value = False
                marked += 1
        print(f"  {sn}: marked {marked} rows crawlable=FALSE")
    wb.save(XLSX)
    print(f"Saved {XLSX}")

if __name__ == "__main__":
    main()
